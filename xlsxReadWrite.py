import openpyxl

class xlsxUtils():

    def __init__(self, path, sheetName):
        self.path = path
        self.sheetName = sheetName
        self.workBook = object
        self.sheet = object
        self.sheetMaxRow = 0
        self.sheetMaxColumn = 0
        self.headerDict = {}
        self.headerDataIndex = {}
        self.initializeSheet()
    
    def initializeSheet(self):

        #Initialize WorkBook.
        self.workBook = openpyxl.load_workbook(self.path)
        
        #Check if sheetname exists in WorkBook.
        if (self.sheetName not in self.workBook.sheetnames):
            errorMessage = "SheetName '%s' not found in WorkBook" % self.sheetName
            raise Exception(errorMessage)
        
        #Initialize Sheet based on SheetName.
        for sheet in self.workBook.worksheets:
            if sheet.title == self.sheetName:
                self.sheet = sheet
                break
        else:
            errorMessage = "SheetName '%s' could not be initialized" % self.sheetName
            raise Exception(errorMessage)
        
        #Define max Column.
        self.__defineMaxColumn()

        #Define max Row.
        self.__defineMaxRow()

        #Create Dict with header data.
        #Create Dict with index of header data.
        for headerColumn in range(1, self.sheetMaxColumn + 1):
            self.headerDict[self.sheet.cell(1, headerColumn).value] = ""
            self.headerDataIndex[self.sheet.cell(1, headerColumn).value] = headerColumn

    def returnData(self):

        #Instantiate list with Dicts.
        sheetData = []
        
        #For each row, add data to a rowDict then append to a list of rows.
        for row in range(2, self.sheetMaxRow + 1):

            #Copy the dict of header data.
            rowDict = self.headerDict.copy()

            for column in range(1, self.sheetMaxColumn + 1):

                #If cell value is not None, add it to the row data dict.
                if (self.sheet.cell(row, column).value != None):
                    rowDict[self.sheet.cell(1, column).value] = self.sheet.cell(row, column).value

            sheetData.append(rowDict)
        
        return sheetData

    def insertNewRow(self, dataDict):

        #Get the number of the last row and add 1.
        newRow = self.sheetMaxRow + 1

        #Iterate through dataDict.
        for key, value in dataDict.items():
            
            #Check to see if all headers exist, if not raise Exception.
            if self.headerDataIndex.get(key) == None:
                errorMessage = "Couldnt find Index for Header '%s'" % key
                raise Exception(errorMessage)

            #Assign the value given in the data dict to the correspondent cell.
            self.sheet.cell(newRow, self.headerDataIndex.get(key), value)

            #Save WorkBook
            self.workBook.save(self.path)

        #Initialize the sheet again.
        self.initializeSheet()
    
    def selectValues(self, paramsDict, responseList):

        #Check if 'paramsDict' and 'responseList' are not empty
        if ((len(paramsDict) == 0) or (len(responseList) == 0)):
            errorMessage = "Invalid params. The paramsDict and responseList can not be empty"
            raise Exception(errorMessage)

        #Instatiate results list.
        resultsList = []

        #Iterate through dataDict.
        for key in paramsDict.keys():

            #Check to see if all headers exist, if not raise Exception.
            if self.headerDataIndex.get(key) == None:
                errorMessage = "Couldnt find Index for Header '%s'" % key
                raise Exception(errorMessage)

        #Iterate through responseList.
        for reponse in responseList:  

            #Check to see if all headers exist, if not raise Exception.
            if self.headerDataIndex.get(reponse) == None:
                errorMessage = "Couldnt find Index for Header '%s'" % key
                raise Exception(errorMessage)


        #For each row, check if all params match, if they do, add response to result list.
        for row in range(2, self.sheetMaxRow + 1):

            #Instantiate list of booleans to check if every param in paramsDict was find.
            controlListOfBooleans = []

            #Append boolean checking if the value of the cell match the param given.
            for key, value in paramsDict.items():
                controlListOfBooleans.append(self.sheet.cell(row, self.headerDataIndex.get(key)).value == value)
            
            #If all booleans are True, all the params match with cell value,
            #thus appending the values of the params given in the response list as the return.
            if (False not in controlListOfBooleans):
                resultDict = {}
                for response in responseList:
                    resultDict[response] = self.sheet.cell(row, self.headerDataIndex.get(response)).value
                
                resultsList.append(resultDict)
            
        #If resultsList has more than 1 resultDict, the query found more than 1 result,
        #thus raising an Exception.
        #If resultsList doesnt have any resultDict, a Exception will be raised.
        if len(resultsList) != 1:
            errorMessage = "%s results were found in the query, consider refining it." % str(len(resultsList))
            raise Exception(errorMessage)

        return resultsList[0]
    
    def updateValues(self, paramsDict, valuesToBeUpdated):

        #Check if 'paramsDict' and 'responseList' are not empty
        if ((len(paramsDict) == 0) or (len(valuesToBeUpdated) == 0)):
            errorMessage = "Invalid params. The paramsDict and responseList can not be empty"
            raise Exception(errorMessage)
        
        #Instatiate counters.
        rowsAffected = 0
        cellsUpdated = 0

        #Iterate through dataDict keys.
        for key in paramsDict.keys():

            #Check to see if all headers exist, if not raise Exception.
            if self.headerDataIndex.get(key) == None:
                errorMessage = "Couldnt find Index for Header '%s'" % key
                raise Exception(errorMessage)

        #Iterate through valuesToBeUpdated keys.
        for key in valuesToBeUpdated.keys():  

            #Check to see if all headers exist, if not raise Exception.
            if self.headerDataIndex.get(key) == None:
                errorMessage = "Couldnt find Index for Header '%s'" % key
                raise Exception(errorMessage)


        #For each row, check if all params match, if they do, update the values based on valuesToBeUpdated Dict.
        for row in range(2, self.sheetMaxRow + 1):

            #Instantiate list of booleans to check if every param in paramsDict was find.
            controlListOfBooleans = []

            #Append boolean checking if the value of the cell match the param given.
            for key, value in paramsDict.items():
                controlListOfBooleans.append(self.sheet.cell(row, self.headerDataIndex.get(key)).value == value)
            
            #If all booleans are True, all the params match with cell value,
            #thus updating the values of the params based on valuesToBeUpdated Dict.
            if (False not in controlListOfBooleans):
                rowsAffected += 1
                for key, value in valuesToBeUpdated.items():
                    self.sheet.cell(row, self.headerDataIndex.get(key), value)
                    cellsUpdated += 1

                #Save WorkBook.
                self.workBook.save(self.path)
                
            
        #If resultsList has more than 1 resultDict, the query found more than 1 result,
        #thus raising an Exception.
        #If resultsList doesnt have any resultDict, a Exception will be raised.
        print("%s rows affected, %s cells updated" % (str(rowsAffected), str(cellsUpdated)))
    
    def __defineMaxColumn(self):

        value = ""
        column = 1
        while value != None:
            value = self.sheet.cell(1,column).value
            column += 1
        
        self.sheetMaxColumn = column - 2
    
    def __defineMaxRow(self):

        value = ""
        row = 1
        while value != None:
            value = self.sheet.cell(row,1).value
            row += 1
        
        self.sheetMaxRow = row - 2